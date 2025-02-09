from models import *
import pandas as pd
import openpyxl
from columnar import columnar
import calendar
from calendar import monthrange
import subprocess
import xlwings as xw

def print_calendar(calendar):
    """
    Print the calendar in a digestible 5-week grid format.
    Each day shows who is scheduled for each of the 4 shifts.
    """
    days_per_week = 7
    column_width = 10  # Fixed width for each column

    print("\nSchedule:")
    print("=" * (column_width * days_per_week))

    # Calculate the weekday of the first date in the calendar (0 = Sunday, 6 = Saturday)
    if not calendar:
        print("No days in the calendar.")
        return

    first_day_weekday = calendar[0].date.weekday()
    first_day_weekday = (first_day_weekday + 1) % 7  # Convert Python's weekday (0=Monday) to Sunday-start

    # Add an offset for the first week
    week_offset = [" " * column_width] * first_day_weekday

    # Group days into weeks
    weeks = []
    current_week = week_offset

    for day in calendar:
        current_week.append(day)
        if len(current_week) == days_per_week:
            weeks.append(current_week)
            current_week = []

    if current_week:
        # Fill the last week with empty spaces if necessary
        current_week.extend([" " * column_width] * (days_per_week - len(current_week)))
        weeks.append(current_week)

    # Print each week
    for week in weeks:
        # Print dates as headers
        header_row = "".join(
            f"{(day.date.strftime('%b %d') if isinstance(day, CalDay) else day):^{column_width}}"
            for day in week
        )
        print(header_row)

        # Print scheduled doctors for each shift
        for shift in ["s1", "s2", "s3", "s4"]:  # Shift 1 to Shift 4
            shift_row = "".join(
                f"{(day.shifts[shift].name if isinstance(day, CalDay) and day.shifts[shift] else '-----'):^{column_width}}"
                if isinstance(day, CalDay)
                else f"{day:^{column_width}}"
                for day in week
            )
            print(shift_row)

        # Add a blank line between weeks
        print("\n")

def load_month_and_year(filepath):
    """
    Load the month and year for scheduling from the "Color" sheet in the Excel file.

    Args:
        filepath (str): Path to the Excel file.

    Returns:
        tuple: (month, year) as integers.
    """
    # Load the "Color" sheet
    color_sheet = pd.read_excel(filepath, sheet_name="Color", header=None)
    
    # Month is in cell L2 (row 1, column 11 in zero-based indexing)
    month = int(color_sheet.iloc[1, 11])
    
    # Year is in cell L3 (row 2, column 11 in zero-based indexing)
    year = int(color_sheet.iloc[2, 11])
    
    return month, year

def load_doctor_inputs(filepath):
    """
    Load doctor inputs from the Excel file and initialize Doctor objects.

    Args:
        filepath (str): Path to the Excel file.

    Returns:
        list: List of Doctor objects.
    """
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook["Doctor Inputs"]

    doctors = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip rows without a name
            continue

        name = row[0].strip()
        doc_type = row[2]
        min_shifts, max_shifts = map(int, row[5].split(","))
        shift_prefs = list(map(int, row[6].split(",")))
        flip_shifts = bool(row[7])

        # Create a Doctor object
        doctor = Doctor(
            name=name,
            days_off=[],  # Temporarily empty, updated in the next step
            shift_prefs=shift_prefs,
            min_shifts=min_shifts,
            max_shifts=max_shifts,
            flip_shifts=flip_shifts,
            doc_type=doc_type
        )
        doctors.append(doctor)

    return doctors

def load_shifts_requested_off(filepath, doctors, schedule_month, schedule_year):
    """
    Load shifts requested off and update the days_off attribute for each Doctor.

    Args:
        filepath (str): Path to the Excel file.
        doctors (list): List of Doctor objects.
        schedule_month (int): The month of the schedule (1-12).
        schedule_year (int): The year of the schedule.
    """
    wb = xw.Book(filepath)
    sheet = wb.sheets["Scheduling Worksheet"]
    df = sheet.used_range.value  # Read with live calculation
    worksheet = pd.DataFrame(df)

    # Calculate the number of days in the given month
    num_days = calendar.monthrange(schedule_year, schedule_month)[1]

    for doctor in doctors:
        doctor_row = worksheet[worksheet[0] == doctor.name]
        if doctor_row.empty:
            continue

        flip_shifts = doctor.flip_shifts
        days_off = []

        # Loop through only the actual days in the month
        for col_idx in range(5, 5 + num_days):  # Columns start at 5 for the 1st day
            day = col_idx - 4  # Calculate the day (1-indexed)
            value = doctor_row.iloc[0, col_idx]

            # Treat NaN as False
            is_day_off = pd.isna(value) or value == 0

            # Adjust logic for flip_shifts and non-flip_shifts
            if (flip_shifts and is_day_off) or (not flip_shifts and not is_day_off):
                days_off.append(date(schedule_year, schedule_month, day))

        # Assign processed days_off to the doctor
        doctor.days_off = set(days_off)

def load_previous_month_shifts(filepath, doctors, schedule_month, schedule_year):
    last_4_days = get_last_days_of_previous_month(schedule_month, schedule_year)

    wb = xw.Book(filepath)
    sheet = wb.sheets["Scheduling Worksheet"]
    df = sheet.used_range.value  # Read with live calculation
    worksheet = pd.DataFrame(df)

    # Extract values and filter out invalid rows
    rows = list(worksheet.values)
    data_rows = [row for row in rows if row[0] and isinstance(row[0], str) and row[0].strip()]

    for doctor in doctors:
        # Find the row for the doctor
        doctor_row = next((row for row in data_rows if row[0].strip() == doctor.name), None)

        if doctor_row is None or len(doctor_row) == 0:
            #print(f"WARNING: No match found for Doctor {doctor.name} in the worksheet.")
            continue

        previous_month_shifts = []
        for col_idx, day in zip(range(1, 5), last_4_days):  # Adjust column indices if needed
            shift_type = doctor_row[col_idx] if len(doctor_row) > col_idx else None

            # Ensure shift_type is an integer if valid, otherwise None
            if pd.isna(shift_type) or shift_type is None:
                shift_type = None
            else:
                shift_type = int(shift_type) if isinstance(shift_type, (int, float)) else None

            if shift_type is not None:
                previous_month_shifts.append((day, shift_type))

        doctor.previous_month_shifts = previous_month_shifts

def get_last_days_of_previous_month(schedule_month, schedule_year):
    if schedule_month == 1:
        prev_month = 12
        prev_year = schedule_year - 1
    else:
        prev_month = schedule_month - 1
        prev_year = schedule_year

    num_days_prev_month = calendar.monthrange(prev_year, prev_month)[1]
    last_4_days = [date(prev_year, prev_month, num_days_prev_month - i) for i in range(4)][::-1]
    return last_4_days


def print_doctor_info(doctors):
    for doctor in doctors:
        print(f"Doctor {doctor.name}:")
        print(f"  - Days Off: {', '.join(str(day) for day in sorted(doctor.days_off)) if doctor.days_off else 'None'}")
        print(f"  - Shift Preferences: {doctor.shift_prefs}")
        print(f"  - Min Shifts: {doctor.min_shifts}, Max Shifts: {doctor.max_shifts}")
        
        # Format Previous Month Shifts
        if doctor.previous_month_shifts:
            formatted_shifts = ', '.join(f"Day {day}: Shift {shift}" for day, shift in doctor.previous_month_shifts)
            print(f"  - Previous Month Shifts: {formatted_shifts}")
        else:
            print(f"  - Previous Month Shifts: None")
        
        print(f"  - Total Shifts Scheduled: {doctor.total_shifts}, Consecutive Shifts: {doctor.consecutive_shifts}")
        print(f"  - Night Shifts: {doctor.night_shifts}, Weekend Shifts: {doctor.weekend_shifts}")
        print("")

import openpyxl
from datetime import datetime

def write_scheduled_shifts(filepath, calendar, schedule_month, schedule_year):
    """
    Write scheduled shifts back into the "Color" tab of the Excel file.

    Args:
        filepath (str): Path to the Excel file.
        calendar (list): List of CalDay objects containing shift assignments.
        schedule_month (int): The month of the schedule (1-12).
        schedule_year (int): The year of the schedule.
    """
    # Load the workbook and "Color" sheet
    workbook = openpyxl.load_workbook(filepath)
    color_sheet = workbook["Color"]

    # Define rows corresponding to each week's dates
    date_rows = [4, 11, 18, 25, 32, 39]

    # Determine the starting column for the 1st day of the month
    first_day_of_month = datetime(schedule_year, schedule_month, 1).date()

    weekday_to_column = {
        6: 2,  # Sunday -> Column B
        0: 3,  # Monday -> Column C
        1: 4,  # Tuesday -> Column D
        2: 5,  # Wednesday -> Column E
        3: 6,  # Thursday -> Column F
        4: 7,  # Friday -> Column G
        5: 8,  # Saturday -> Column H
    }
    start_column = weekday_to_column[first_day_of_month.weekday()]

    # Loop through each CalDay object in the calendar
    for cal_day in calendar:
        day_offset = (cal_day.date - first_day_of_month).days

        # Calculate the column for the current day (with wrapping)
        column = (start_column + day_offset - 2) % 7 + 2

        # Determine the base row for the week of the date
        week_index = (start_column + day_offset - 2) // 7
        if week_index < len(date_rows):
            week_base_row = date_rows[week_index]
        else:
            raise ValueError(f"Invalid date {cal_day.date} for scheduling")

        # Write scheduled shifts for the day
        for shift_type, doctor in cal_day.shifts.items():
            if doctor:
                # Convert shift type (e.g., "s1", "s2") to row offset
                shift_number = int(shift_type[1])
                shift_row = week_base_row + shift_number

                # Write the doctor's name into the cell
                cell = color_sheet.cell(row=shift_row, column=column)
                cell.value = doctor.name

    # Save the workbook
    workbook.save(filepath)
    open_excel_file(filepath)

def read_manual_shift4_assignments(filepath, calendar, doctors, schedule_month, schedule_year, scheduler):
    """
    Read manually assigned shift 4 schedules from Excel and update the calendar.

    Args:
        filepath (str): Path to the Excel file.
        calendar (list): List of CalDay objects representing the schedule.
        doctors (list): List of Doctor objects.
        schedule_month (int): The month of the schedule (1-12).
        schedule_year (int): The year of the schedule.
        scheduler (Scheduler): The scheduler instance to assign shifts.
    """
    # Load the workbook and select the "Color" sheet
    workbook = openpyxl.load_workbook(filepath, data_only=True)
    color_sheet = workbook["Color"]

    # Define rows corresponding to each week's dates
    date_rows = [4, 11, 18, 25, 32, 39]  # Start rows for each week

    # Determine the starting column for the 1st day of the month
    first_day_of_month = datetime(schedule_year, schedule_month, 1)
    weekday_to_column = {
        6: 2,  # Sunday -> Column B
        0: 3,  # Monday -> Column C
        1: 4,  # Tuesday -> Column D
        2: 5,  # Wednesday -> Column E
        3: 6,  # Thursday -> Column F
        4: 7,  # Friday -> Column G
        5: 8,  # Saturday -> Column H
    }
    start_column = weekday_to_column[first_day_of_month.weekday()]

    # Reset shift counters before processing updated assignments
    for doc in doctors:
        doc.total_shifts = 0
        doc.night_shifts = 0
        doc.weekend_shifts = 0
        doc.last_shift_date = datetime.min.date()

    # Clear any previous Shift 4 assignments before applying new ones
    for cal_day in calendar:
        cal_day.shifts["s4"] = None  # Reset previous assignments

    # Iterate over each week
    for week_index, week_base_row in enumerate(date_rows):
        for day_offset in range(7):  # 7 days in a week
            # Calculate column
            column = start_column + day_offset
            if column == 9:
                week_base_row += 7
            if column > 8:  # If past column H, reset to B and increment row
                column = 2 + (column - 9)

            # Calculate actual shift date
            shift_date = first_day_of_month + timedelta(days=week_index * 7 + day_offset)
            shift_date = shift_date.date()  # Convert to datetime.date

            # Find corresponding calendar day
            cal_day = next((day for day in calendar if day.date == shift_date), None)

            # Debugging: Print column/row mapping
            #print(f"DEBUG: Processing {shift_date.strftime('%b %d')} | Row: {week_base_row + 4}, Column: {column} | CalDay: {cal_day}")

            # Ensure the date is within the current month
            if shift_date.month != schedule_month:
                continue

            # If cal_day is None, print an error and continue
            if not cal_day:
                print(f"WARNING: No matching calendar day found for {shift_date}. Skipping.")
                continue

            # Determine the row where Shift 4 assignments are stored
            shift4_row = week_base_row + 4  # Shift 4 is always row 4 below the date row

            # Read doctor name from Excel
            cell_value = color_sheet.cell(row=shift4_row, column=column).value

            # Find the corresponding doctor object
            assigned_doctor = next((doc for doc in doctors if doc.name == cell_value), None)
            if assigned_doctor is not None:
                scheduler.assign_shift(cal_day, assigned_doctor, "s4")
            else:
                print(f"WARNING: No matching doctor found for {shift_date.strftime('%b %d')}!")

    # Debug: Print final state of shift 4 assignments
    #for cal_day in calendar:
        #print(f"DEBUG CHECK: {cal_day.date.strftime('%b %d')} Shift 4: {cal_day.shifts['s4'].name if cal_day.shifts['s4'] else 'None'}")

def open_excel_file(filepath):
    """
    Opens the specified Excel file using the default application.
    
    Args:
        filepath (str): The full path to the Excel file.
    """
    subprocess.run(["open", filepath])

def debug_print_doctor_shifts(doctors):
    """
    Print out the total shifts, night shifts, weekend shifts, and consecutive shifts for each doctor
    in a readable format.
    
    Args:
        doctors (list): List of Doctor objects.
    """
    print("\n=== DEBUG: Doctor Shift Summary ===")
    print(f"{'Doctor':<10} | {'Total Shifts':<12} | {'Night Shifts':<12} | {'Weekend Shifts':<15} | {'Consecutive Days'}")
    print("-" * 70)

    for doctor in doctors:
        print(f"{doctor.name:<10} | {doctor.total_shifts:<12} | {doctor.night_shifts:<12} | {doctor.weekend_shifts:<15} | {doctor.consecutive_shifts}")

    print("=" * 70 + "\n")

def clear_scheduled_shifts(filepath):
    """
    Clears all scheduled shifts (S1, S2, S3) from the Excel output file while keeping S4 intact.

    Args:
        filepath (str): Path to the Excel file.
    """
    wb = xw.Book(filepath)
    sheet = wb.sheets["Color"]  # Adjust sheet name if needed

    # Define the shift row groups: S1, S2, S3 are in these row sets, shifting down every 7 rows
    shift_row_groups = [(5, 6, 7), (12, 13, 14), (19, 20, 21), (26, 27, 28), (33, 34, 35), (40, 41, 42)]

    # Shift columns: B to H (Excel column 2 to 8)
    start_col, end_col = 2, 8

    # Loop through all shift row groups and clear the cells
    for s1, s2, s3 in shift_row_groups:
        for row in [s1, s2, s3]:  # Clear only S1, S2, S3
            sheet.range((row, start_col), (row, end_col)).value = None

    wb.save(filepath)