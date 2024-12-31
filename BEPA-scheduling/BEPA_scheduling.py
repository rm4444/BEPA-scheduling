import calendar
import array
import string
import sys
import xlrd
import openpyxl
import re
import numpy as np
from columnar import columnar
import math
import random
import os

class Doctor:
    def __init__(self,name,days_off,OneSO,TwoSO,ThreeSO,FourSO,shift_prefs,min_shifts,max_shifts):
        self.name = name
        self.days_off = days_off
        self.OneSO = OneSO
        self.TwoSO = TwoSO
        self.ThreeSO = ThreeSO
        self.FourSO = FourSO
        self.shift_prefs = shift_prefs
        self.min_shifts = min_shifts
        self.max_shifts = max_shifts
        self.shifts = 0
        self.weekends = 0
        self.consec_shifts = 0
        self.four_shifts = 0
        self.lastSixDays = 0
        self.lastNineDays = 0

class CalDay:
    def __init__(self,date):
        self.date = date
        self.s1 = Doctor("--",[],[],[],[],[],[],0,0)
        self.s2 = Doctor("--",[],[],[],[],[],[],0,0)
        self.s3 = Doctor("--",[],[],[],[],[],[],0,0)
        self.s4 = Doctor("--",[],[],[],[],[],[],0,0)
        self.weekend = False

class Gap:
    def __init__(self,start_date,length):
        self.start_date = start_date
        self.length = length
        self.remaining = length
        self.filled = False
        self.docs = []

def read_input(fname): #read Excel and create Doctor objects
    docs = []
    wb = xlrd.open_workbook(fname)
    sheet = wb.sheet_by_name("Input")
    month = int(sheet.cell_value(1,1))
    year = int(sheet.cell_value(1,2))
    n = sheet.nrows
    num_days = calendar.monthrange(year,month)[1]
    for i in range(2, n):
        name = sheet.cell_value(i,4).strip().upper()
        min_shifts = list(map(int, re.split(r'\s|,\s',sheet.cell_value(i,5).strip())))[0]
        max_shifts = list(map(int, re.split(r'\s|,\s',sheet.cell_value(i,5).strip())))[1]
        if sheet.cell_value(i,6):
            if sheet.cell_value(i,15):
                if len(sheet.cell_value(i,6).strip()) <= 1:
                    days_on = [int(sheet.cell_value(i,6).strip())]
                else:
                    days_on = list(map(int, re.split(r'\s|,\s',sheet.cell_value(i,6).strip())))
                days_off = [day for day in range(1,num_days+1) if day not in days_on]
            else:
                if len(sheet.cell_value(i,6).strip()) <= 1:
                    days_off = [int(sheet.cell_value(i,6).strip())]
                else:
                    days_off = list(map(int, re.split(r'\s|,\s',sheet.cell_value(i,6).strip())))
        else:
            if sheet.cell_value(i,15):
                days_off = [day for day in range(1,num_days+1)]
            else:
                days_off = []
        if sheet.cell_value(i,7):
            if isinstance(sheet.cell_value(i,7),float):
                OneSO = [int(sheet.cell_value(i,7))]
            else:
                OneSO = list(map(int, re.split(r'\s|,\s',sheet.cell_value(i,7).strip())))
        else:
            OneSO = []
        if sheet.cell_value(i,8):
            if isinstance(sheet.cell_value(i,8),float):
                TwoSO = [int(sheet.cell_value(i,8))]
            else:
                TwoSO = list(map(int, re.split(r'\s|,\s',sheet.cell_value(i,8).strip())))
        else:
            TwoSO = []
        if sheet.cell_value(i,9):
            if isinstance(sheet.cell_value(i,9),float):
                ThreeSO = [int(sheet.cell_value(i,9))]
            else:
                ThreeSO = list(map(int, re.split(r'\s|,\s',sheet.cell_value(i,9).strip())))
        else:
            ThreeSO = []
        if sheet.cell_value(i,10):
            if isinstance(sheet.cell_value(i,10),float):
                FourSO = [int(sheet.cell_value(i,10))]
            else:
                FourSO = list(map(int, re.split(r'\s|,\s',sheet.cell_value(i,10).strip())))
        else:
            FourSO = []
        shift_prefs = list(map(int, re.split(r'\s|,\s',sheet.cell_value(i,11).strip())))
        doc = Doctor(name, days_off, OneSO, TwoSO, ThreeSO, FourSO, shift_prefs, min_shifts, max_shifts)
        docs.append(doc)
    cal1 = CalDay(-3)
    cal2 = CalDay(-2)
    cal3 = CalDay(-1)
    startCal(docs, sheet, n, cal1,cal2,cal3)
        
    return docs, month, year, cal1, cal2, cal3

def startCal(docs, sheet, n, cal1, cal2, cal3):  #read in and create Calendar days for last three days of previous month
    for i in range(2, n):
        if sheet.cell_value(i,12):
            if int(sheet.cell_value(i,12)) == 1:
                cal1.s1 = docs[i-2]
            elif int(sheet.cell_value(i,12)) == 2:
                cal1.s2 = docs[i-2]
            elif int(sheet.cell_value(i,12)) == 3:
                cal1.s3 = docs[i-2]
            elif int(sheet.cell_value(i,12)) == 4:
                cal1.s4 = docs[i-2]
    for i in range(2, n):
        if sheet.cell_value(i,13):
            if int(sheet.cell_value(i,13)) == 1:
                cal2.s1 = docs[i-2]
            elif int(sheet.cell_value(i,13)) == 2:
                cal2.s2 = docs[i-2]
            elif int(sheet.cell_value(i,13)) == 3:
                cal2.s3 = docs[i-2]
            elif int(sheet.cell_value(i,13)) == 4:
                cal2.s4 = docs[i-2]
    for i in range(2, n):
        if sheet.cell_value(i,14):
            if int(sheet.cell_value(i,14)) == 1:
                cal3.s1 = docs[i-2]
            elif int(sheet.cell_value(i,14)) == 2:
                cal3.s2 = docs[i-2]
            elif int(sheet.cell_value(i,14)) == 3:
                cal3.s3 = docs[i-2]
            elif int(sheet.cell_value(i,14)) == 4:
                cal3.s4 = docs[i-2]

def printCal(docs,cal,year,month):
    num_days = calendar.monthrange(year,month)[1]
    firstDay = calendar.monthrange(year,month)[0]
    if month > 1:
        prevMonthDays = calendar.monthrange(year,month-1)[1]
    else:
        prevMonthDays = calendar.monthrange(year-1,12)[1]
    if firstDay >= 2:
        headers = [""]*(firstDay-2) +[str(prevMonthDays-2),str(prevMonthDays-1),str(prevMonthDays)] + list(map(str,range(1,num_days+1)))
        shift1 = [""]*(firstDay-2)
        shift2 = [""]*(firstDay-2)
        shift3 = [""]*(firstDay-2)
        shift4 = [""]*(firstDay-2)
    else:
        headers = [""]*(firstDay+5) +[str(prevMonthDays-2),str(prevMonthDays-1),str(prevMonthDays)] + list(map(str,range(1,num_days+1)))
        shift1 = [""]*(firstDay+5)
        shift2 = [""]*(firstDay+5)
        shift3 = [""]*(firstDay+5)
        shift4 = [""]*(firstDay+5)
    for day in cal:
        shift1.append(day.s1.name)
        shift2.append(day.s2.name)
        shift3.append(day.s3.name)
        shift4.append(day.s4.name)
    if firstDay >= 2:
        printDays = num_days + firstDay - 2 + 3
    else:
        printDays = num_days + firstDay + 3 + 4
    printRange = math.ceil(printDays/7)
    printRange = printRange * 7
    for i in range(0,printRange,7):
        start = 0 + i
        stop = min(7 + i, len(headers))
        heads = headers[start:stop]
        data = [shift1[start:stop],shift2[start:stop],shift3[start:stop],shift4[start:stop]]
        print(columnar(data,heads,no_borders=True))
        print()
    for doc in docs:
        print(doc.name + ": " + str(doc.shifts+doc.four_shifts) + " total shifts and " + str(doc.weekends) + " weekend shifts.")
    print()

def last3days(docs,cal):  #sets consec_shifts for docs based on last three days of work
    for i in range(3):
        cal[i].s1.consec_shifts += 1
        cal[i].s2.consec_shifts += 1
        cal[i].s3.consec_shifts += 1
        cal[i].s4.consec_shifts += 1
        worked = [cal[i].s1,cal[i].s2,cal[i].s3,cal[i].s4]
        notWorked = set(docs) - set(worked)
        for doc in notWorked:
            doc.consec_shifts = 0
        
        
def buildCal(docs, cal, num_days, year, month, fname):
    os.system("cls||clear")
    print()
    last3days(docs,cal)
    inp = input("Type 'Y' to make a new schedule from scratch, and type 'N' to only schedule the day shifts: ").strip().upper()
    while inp != "Y" and inp != "N":
        inp = input("Geoff, you ignorant slut. Type 'Y' to make a new schedule from scratch, and type 'N' to only schedule the day shifts: ").strip().upper()
    if inp == "Y":
        schedule4Shifts(docs,cal,num_days) 
        #check4Shifts(docs,cal,year,month)
        #printCal(docs,cal,year,month)
        exportCal(fname, cal, year, month)
        os.system('cls||clear')
        cmd = "start EXCEL.EXE " + fname
        os.system(cmd)
        print()
        inp = input("When you're done setting the night shifts, save and close out of the Excel document. Press enter when you are ready to continue, you abominable nincompoop: ")

    readInFourShifts(fname, docs, cal, year, month)
    for doc in docs:
        doc.four_shifts = 0
        doc.weekends = 0
    for i in range(3,len(cal)):
        cal[i].s4.four_shifts += 1
        if cal[i].weekend == True:
            cal[i].s4.weekends += 1
        
    for doc in docs:
        doc.max_shifts -= doc.four_shifts
        doc.min_shifts -= doc.four_shifts
        doc.consec_shifts = 0
    for i in range(3):
        cal[i].s1.consec_shifts += 1
        cal[i].s2.consec_shifts += 1
        cal[i].s3.consec_shifts += 1
        cal[i].s4.consec_shifts += 1
        worked = [cal[i].s1,cal[i].s2,cal[i].s3,cal[i].s4]
        notWorked = set(docs) - set(worked)
        for doc in notWorked:
            doc.consec_shifts = 0
    
    for i in range(num_days):
        date = i + 1
        index = i + 3
        calcLastSix(docs,cal,index)
        calcLastNine(docs,cal,index)
        schedule1Shift(docs,cal,index)
        schedule2Shift(docs,cal,index)
        schedule3Shift(docs,cal,index)
        docsWorking = [cal[index].s1,cal[index].s2,cal[index].s3,cal[index].s4]
        for doc in docs:
            if doc in docsWorking:
                doc.consec_shifts += 1
            else:
                doc.consec_shifts = 0
    #printCal(docs,cal,year,month)

def calcLastSix(docs, cal, index):
    for doc in docs:
        doc.lastSixDays = 0
    for i in range(max(0,index - 6),index):
        cal[i].s1.lastSixDays += 1
        cal[i].s2.lastSixDays += 1
        cal[i].s3.lastSixDays += 1
        cal[i].s4.lastSixDays += 1

def calcLastNine(docs, cal, index):
    for doc in docs:
        doc.lastSixDays = 0
    for i in range(max(0,index - 9),index):
        cal[i].s1.lastSixDays += 1
        cal[i].s2.lastSixDays += 1
        cal[i].s3.lastSixDays += 1
        cal[i].s4.lastSixDays += 1

def schedule1Shift(docs,cal,index):
    date = index - 2
    P0 = docs[:]
    for doc in docs: #if doc scheduled a day off, or worked a later shift the day before, or have worked 4 consecutive shifts, or have worked over their max shifts, remove them
        if doc.name == "ADAL" or doc.name == "COL" or doc.name == "PAT" or date in doc.days_off or date in doc.OneSO or cal[index].s4 == doc or cal[index - 1].s2 == doc or cal[index - 1].s3 == doc or cal[index - 1].s4 == doc or doc.consec_shifts >= 4 or doc.shifts >= doc.max_shifts or (cal[min(index+1,len(cal)-1)].s4 == doc and doc.consec_shifts != 0) or cal[index - 2].s4 == doc or (doc.name == "HRA" and doc.consec_shifts >=3):
            P0.remove(doc)
    P1 = [doc for doc in P0 if doc.consec_shifts <= 4 and doc.shifts < doc.min_shifts and doc.lastSixDays < 5 and doc.lastNineDays < 7]
    P2 = [doc for doc in P0 if doc not in P1 and doc.lastSixDays < 5 and doc.lastNineDays < 7]
    P3 = [doc for doc in P0 if doc not in P1 and doc not in P2]
    if len(P1) > 1:
        P1.sort(key = lambda x: x.min_shifts, reverse = True)
        P1.sort(key = lambda x: x.shifts/max(x.min_shifts,1))
        if cal[index].weekend == True:
            P1.sort(key = lambda x: x.weekends)
        P1.sort(key = lambda x: x.shift_prefs[0], reverse = True)

    if len(P1) >= 1:
        schedDoc = P1[0]
    
    elif len(P2) >= 1:
        P2.sort(key = lambda x: x.min_shifts, reverse = True)
        P2.sort(key = lambda x: x.shifts/max(x.min_shifts,1))
        if cal[index].weekend == True:
            P2.sort(key = lambda x: x.weekends)
        P2.sort(key = lambda x: x.shift_prefs[0], reverse = True)
        schedDoc = P2[0]

    elif len(P3) >= 1:
        P3.sort(key = lambda x: x.min_shifts, reverse = True)
        P3.sort(key = lambda x: x.shifts/max(x.min_shifts,1))
        if cal[index].weekend == True:
            P3.sort(key = lambda x: x.weekends)
        P3.sort(key = lambda x: x.shift_prefs[0], reverse = True)
        schedDoc = P3[0]

    else: 
        schedDoc = Doctor("--",[],[],[],[],[],[],0,0)
    """ print(str(date) + " shift 1: " + schedDoc.name)
    print([doc.name + " " + str(doc.shifts) + " " + str(doc.min_shifts) + " " + str(doc.lastSixDays) + " " + str(doc.lastNineDays) for doc in P1])
    print([doc.name + " " + str(doc.shifts) + " " + str(doc.min_shifts) + " " + str(doc.lastSixDays) + " " + str(doc.lastNineDays) for doc in P2])
    print([doc.name + " " + str(doc.shifts) + " " + str(doc.min_shifts) + " " + str(doc.lastSixDays) + " " + str(doc.lastNineDays) for doc in P3]) """
    cal[index].s1 = schedDoc
    schedDoc.shifts += 1
    if cal[index].weekend == True:
        schedDoc.weekends += 1

def schedule2Shift(docs,cal,index):
    date = index - 2
    P0 = docs[:]
    switchROD = False
    switchHRA = False
    for doc in docs: #if doc scheduled a day off, or worked a later shift the day before, or have worked 4 consecutive shifts, or have worked over their max shifts, remove them
        if doc.name == "ROTT" or doc.name == "PAT" or date in doc.days_off or date in doc.TwoSO or cal[index].s4 == doc or cal[index].s1 == doc or cal[index - 1].s3 == doc or cal[index - 1].s4 == doc or doc.consec_shifts >= 4 or doc.shifts >= doc.max_shifts or (cal[min(index+1,len(cal)-1)].s4 == doc and doc.consec_shifts != 0) or cal[index - 2].s4 == doc or (doc.name == "HRA" and doc.consec_shifts >=3):
            P0.remove(doc)
    P1 = [doc for doc in P0 if doc.consec_shifts <= 4 and doc.shifts < doc.min_shifts and doc.lastSixDays < 5 and doc.lastNineDays < 7]
    P2 = [doc for doc in P0 if doc not in P1 and doc.lastSixDays < 5 and doc.lastNineDays < 7]
    P3 = [doc for doc in P0 if doc not in P1 and doc not in P2]

    if len(P1) > 1:
        P1.sort(key = lambda x: x.min_shifts, reverse = True)
        P1.sort(key = lambda x: x.shifts/max(x.min_shifts,1))
        if cal[index].weekend == True:
            P1.sort(key = lambda x: x.weekends)
        P1.sort(key = lambda x: x.shift_prefs[1], reverse = True)

    if len(P1) >= 1:
        schedDoc = P1[0]
    
    elif len(P2) >= 1:
        P2.sort(key = lambda x: x.min_shifts, reverse = True)
        P2.sort(key = lambda x: x.shifts/max(x.min_shifts,1))
        if cal[index].weekend == True:
            P2.sort(key = lambda x: x.weekends)
        P2.sort(key = lambda x: x.shift_prefs[1], reverse = True)
        schedDoc = P2[0]

    elif len(P3) >= 1:
        P3.sort(key = lambda x: x.min_shifts, reverse = True)
        P3.sort(key = lambda x: x.shifts/max(x.min_shifts,1))
        if cal[index].weekend == True:
            P3.sort(key = lambda x: x.weekends)
        P3.sort(key = lambda x: x.shift_prefs[1], reverse = True)
        schedDoc = P3[0]

    else:
        schedDoc = Doctor("--",[],[],[],[],[],[],0,0)
    """ print(str(date) + " shift 2: " + schedDoc.name)
    print([doc.name + " " + str(doc.shifts) + " " + str(doc.min_shifts) + " " + str(doc.lastSixDays) + " " + str(doc.lastNineDays) for doc in P1])
    print([doc.name + " " + str(doc.shifts) + " " + str(doc.min_shifts) + " " + str(doc.lastSixDays) + " " + str(doc.lastNineDays) for doc in P2])
    print([doc.name + " " + str(doc.shifts) + " " + str(doc.min_shifts) + " " + str(doc.lastSixDays) + " " + str(doc.lastNineDays) for doc in P3]) """
    cal[index].s2 = schedDoc
    schedDoc.shifts += 1
    if cal[index].weekend == True:
        schedDoc.weekends += 1

def schedule3Shift(docs,cal,index):
    date = index - 2
    P0 = docs[:]
    for doc in docs: #if doc scheduled a day off, or worked a later shift the day before, or have worked 4 consecutive shifts, or have worked over their max shifts, remove them
        if doc.name == "ROTT" or doc.name == "COL" or doc.name == "ADAL" or doc.name == "PAT" or date in doc.days_off or date in doc.ThreeSO or cal[index].s4 == doc or cal[index].s1 == doc or cal[index].s2 == doc or cal[index - 1].s4 == doc or doc.consec_shifts >= 4 or doc.shifts >= doc.max_shifts or (cal[min(index+1,len(cal)-1)].s4 == doc and doc.consec_shifts != 0) or cal[index - 2].s4 == doc or (doc.name == "HRA" and doc.consec_shifts >=3):
            P0.remove(doc)
    if cal[index].s2.name == "ROD" or cal[index].s4.name == "ROD":
        for doc in P0:
            if doc.name == "HRA":
                P0.remove(doc)
    if cal[index].s2.name == "HRA" or cal[index].s4.name == "HRA":
        for doc in P0:
            if doc.name == "ROD":
                P0.remove(doc)            
    P1 = [doc for doc in P0 if doc.consec_shifts <= 4 and doc.shifts < doc.min_shifts and doc.lastSixDays < 5 and doc.lastNineDays < 7]
    P2 = [doc for doc in P0 if doc not in P1 and doc.lastSixDays < 5 and doc.lastNineDays < 7]
    P3 = [doc for doc in P0 if doc not in P1 and doc not in P2]
    if len(P1) > 1:
        P1.sort(key = lambda x: x.min_shifts, reverse = True)
        P1.sort(key = lambda x: x.shifts/max(x.min_shifts,1))
        if cal[index].weekend == True:
            P1.sort(key = lambda x: x.weekends)
        P1.sort(key = lambda x: x.shift_prefs[2], reverse = True)

    if len(P1) >= 1:
        schedDoc = P1[0]
    
    elif len(P2) >= 1:
        P2.sort(key = lambda x: x.min_shifts, reverse = True)
        P2.sort(key = lambda x: x.shifts/max(x.min_shifts,1))
        if cal[index].weekend == True:
            P2.sort(key = lambda x: x.weekends)
        P2.sort(key = lambda x: x.shift_prefs[2], reverse = True)
        schedDoc = P2[0]

    elif len(P3) >= 1:
        P3.sort(key = lambda x: x.min_shifts, reverse = True)
        P3.sort(key = lambda x: x.shifts/max(x.min_shifts,1))
        if cal[index].weekend == True:
            P3.sort(key = lambda x: x.weekends)
        P3.sort(key = lambda x: x.shift_prefs[2], reverse = True)
        schedDoc = P3[0]
    else:
        schedDoc = Doctor("--",[],[],[],[],[],[],0,0)
    """ print(str(date) + " shift 3: " + schedDoc.name)
    print([doc.name + " " + str(doc.shifts) + " " + str(doc.min_shifts) + " " + str(doc.lastSixDays) + " " + str(doc.lastNineDays) for doc in P1])
    print([doc.name + " " + str(doc.shifts) + " " + str(doc.min_shifts) + " " + str(doc.lastSixDays) + " " + str(doc.lastNineDays) for doc in P2])
    print([doc.name + " " + str(doc.shifts) + " " + str(doc.min_shifts) + " " + str(doc.lastSixDays) + " " + str(doc.lastNineDays) for doc in P3])
    print() """
    cal[index].s3 = schedDoc
    schedDoc.shifts += 1
    if cal[index].weekend == True:
        schedDoc.weekends += 1

        

def schedule4Shifts(docs,cal,num_days):
    P1 = [doc for doc in docs if doc.shift_prefs[3] == 9]
    P2 = [doc for doc in docs if doc.shift_prefs[3] == 8]
    P3 = [doc for doc in docs if doc.shift_prefs[3] == 7]
    P4 = [doc for doc in docs if doc.shift_prefs[3] <= 6]
    
    #schedule first few shifts of month based on who worked the end of last month
    date = 1
    os.system("cls||clear")
    print()
    pShifts = input("How many consecutive days was PAT's most recent series of night shifts? ")
    print()
    while(True):
        if pShifts.isdigit():
            pShifts = int(pShifts)
            if pShifts >= 0:
                break
        else:
            pShifts = input("That's not a valid response, jackass. How many consecutive days was PAT's most recent series of night shifts? ")
    if pShifts >= 4:
        minGap = 3
    else:
        minGap = 3

    if cal[2].s4 in P1:
        max_night_shifts = 4
    elif cal[2].s4 in P2:
        max_night_shifts = min(3,minGap)
    elif cal[2].s4 in P3:
        max_night_shifts = 2
    else:
        max_night_shifts = 1
    if cal[2].s4 != P1[0]:
        schedDoc = cal[2].s4
        #print(schedDoc.name + " " + str(schedDoc.consec_shifts) + " " + str(schedDoc.four_shifts))
        while schedDoc.consec_shifts < 3 and schedDoc.four_shifts < max_night_shifts - 1 and date not in schedDoc.FourSO and date not in schedDoc.days_off and date + 1 not in schedDoc.days_off:
            #print(schedDoc.name + " " + str(schedDoc.consec_shifts) + " " + str(schedDoc.four_shifts))
            cal[date + 2].s4 = schedDoc
            schedDoc.consec_shifts += 1
            schedDoc.four_shifts += 1
            if cal[date + 2].weekend == True:
                schedDoc.weekends += 1
            date += 1
    else:
        schedDoc = P1[0]
        while schedDoc.consec_shifts < 4 and date not in schedDoc.FourSO and date not in schedDoc.days_off and date + 1 not in schedDoc.days_off:
            cal[date + 2].s4 = schedDoc
            schedDoc.consec_shifts += 1
            schedDoc.four_shifts += 1
            if cal[date + 2].weekend == True:
                schedDoc.weekends += 1
            date += 1

    #schedule Patton
    if schedDoc == P1[0]:
        date += schedDoc.consec_shifts - 1
    shiftReset = P1[0].four_shifts
    weekendReset = P1[0].weekends
    schedDoc = P1[0]
    for i in range(date,num_days+1):
        index = i + 2
        if cal[index].s4.name == "--":
            lastDayIndex = i + 2
            while cal[lastDayIndex].s4 != P1[0]:  #find index of last day Patton had a shift
                if lastDayIndex == 0:
                    break
                else:
                    lastDayIndex -= 1
            gap = index - lastDayIndex
            #print("date: " + str(i) + " gap: " + str(gap))
            if gap > minGap:
                if list(set([i,i+1,i+2])&set(schedDoc.days_off)) == [] and list(set([i,i+1,i+2])&set(schedDoc.FourSO)) == []:
                    cal[index].s4 = schedDoc
                    if index+1 <= len(cal)-1:
                        cal[index+1].s4 = schedDoc
                        schedDoc.four_shifts += 2
                        if cal[index].weekend == True:
                            schedDoc.weekends += 1
                        if cal[index + 1].weekend == True:
                            schedDoc.weekends += 1
                    minGap = 3
                    if i+3 not in schedDoc.days_off and i+3 not in schedDoc.FourSO:
                        if index+2 <= len(cal)-1:
                            cal[index + 2].s4 = schedDoc
                            schedDoc.four_shifts += 1
                            if cal[index + 2].weekend == True:
                                schedDoc.weekends += 1
                        minGap = 3
                        rand = random.randint(0,1)
                        if i+4 not in schedDoc.days_off and i+4 not in schedDoc.FourSO and i+3 != num_days and rand == 1:
                            if index+3 <= len(cal)-1:
                                cal[index+3].s4 = schedDoc
                                schedDoc.four_shifts += 1
                                if cal[index + 3].weekend == True:
                                    schedDoc.weekends += 1
                                minGap = 3
        #print(str(i) + ": " + str(P1[0].four_shifts))

##    #if Patton does not have enough shifts
##    if P1[0].four_shifts < 0:
##        P1[0].four_shifts = shiftReset
##        P1[0].weekends = weekendReset
##        date -= 1
##        for i in range(date+2,len(cal)):
##            cal[i].s4 = Doctor("--",[],[],[],[],[],[],0,0)
##        minGap = 2
##        for i in range(date,num_days+1):
##            index = i + 2
##            if cal[index].s4.name == "--":
##                lastDayIndex = i + 2
##                while cal[lastDayIndex].s4 != P1[0]:  #find index of last day Patton had a shift
##                    if lastDayIndex == 0:
##                        break
##                    else:
##                        lastDayIndex -= 1
##                gap = index - lastDayIndex
##                if gap > minGap:
##                    if list(set([i,i+1,i+2])&set(schedDoc.days_off)) == [] and list(set([i,i+1,i+2])&set(schedDoc.FourSO)) == []:
##                        cal[index].s4 = schedDoc
##                        if index+1 <= len(cal)-1:
##                            cal[index+1].s4 = schedDoc
##                            schedDoc.four_shifts += 2
##                            if cal[index].weekend == True:
##                                schedDoc.weekends += 1
##                            if cal[index + 1].weekend == True:
##                                schedDoc.weekends += 1
##                        minGap = 2
##                        if i+3 not in schedDoc.days_off and i+3 not in schedDoc.FourSO:
##                            if index+2 <= len(cal)-1:
##                                cal[index + 2].s4 = schedDoc
##                                schedDoc.four_shifts += 1
##                                if cal[index + 2].weekend == True:
##                                    schedDoc.weekends += 1
##                            minGap = 2
##                            if i+4 not in schedDoc.days_off and i+4 not in schedDoc.FourSO and i+3 != num_days:
##                                if index+3 <= len(cal)-1:
##                                    cal[index+3].s4 = schedDoc
##                                    schedDoc.four_shifts += 1
##                                    if cal[index + 3].weekend == True:
##                                        schedDoc.weekends += 1
##                                minGap = 2     
    #Determine which doc worked most recent gap before first gap
    docNames = [doc.name for doc in docs]
    inp = input("Which docs (other than PAT) have worked night shifts in the 7 days leading up to this month? Separate multiple docs with a comma and space: ")
    lastDocsOG = re.split(r'\s|,\s',inp.strip().upper())
    while(True):
        arr = [name for name in lastDocsOG if name not in docNames]
        if len(arr) == 0:
            break
        else:
            inp = input("Error: doctor(s) not recognized. Is this really that hard for you? I'm amazed you've even made it this far. Make sure you spell them the same way as on the input sheet, and separate multiple docs with a comma and a space: ")
            lastDocsOG = re.split(r'\s|,\s',inp.strip().upper())   
    
    #Create GapMap
    gapMap = createGapMap(cal,num_days)
    #print([str(gap.start_date) + ", " + str(gap.length) for gap in gapMap])

    #KED and MIZ
    for i in range(len(gapMap)):
        gap = gapMap[i]
        gapDates = [gap.start_date + j for j in range(gap.length)]
        if gap.length >= 3:
            gapDocs = P2[:]
            for doc in gapDocs:
                if doc.four_shifts > 3:
                    gapDocs.remove(doc)
            gapDocs.sort(key = lambda x: x.four_shifts)
            schedule4Gaps(cal,gap,gapDocs,gapDates,3,lastDocsOG)

    #KED and MIZ and ROD and TOM
    gapMap = createGapMap(cal,num_days)
    #print([str(gap.start_date) + ", " + str(gap.length) for gap in gapMap])

    for i in range(len(gapMap)):
        gap = gapMap[i]
        gapDates = [gap.start_date + j for j in range(gap.length)]
        if gap.length >= 2:
            gapDocs = sorted(P3[:],key = lambda x: x.four_shifts) + sorted(P2[:], key = lambda x: x.four_shifts)
            if gap.length >= 3:
                gapDocs = sorted(P2[:],key = lambda x: x.four_shifts) + sorted(P3[:], key = lambda x: x.four_shifts)
            for doc in gapDocs:
                if doc.four_shifts > 4:
                    gapDocs.remove(doc)
            schedule4Gaps(cal,gap,gapDocs,gapDates,2,lastDocsOG)

    #ROD and TOM
    gapMap = createGapMap(cal,num_days)
    #print([str(gap.start_date) + ", " + str(gap.length) for gap in gapMap])

    for i in range(len(gapMap)):
        gap = gapMap[i]
        gapDates = [gap.start_date + j for j in range(gap.length)]
        gapDocs = P3[:]
        gapDocs.sort(key = lambda x: x.four_shifts)
        schedule4Gaps(cal,gap,gapDocs,gapDates,1,lastDocsOG)

                
def schedule4Gaps(cal,gap,gapDocs,gapDates,min_shifts,lastDocsOG):
    for doc in gapDocs:
        for i in range(len(gapDates)-min_shifts + 1):
            date = gapDates[i]
            index = date + 2
            lastDocs = []
            if index <= 6:
                lastDocs = lastDocsOG[:]
            else:
                for q in range(index - 7, min(index + 7,len(cal))):
                    lastDocs.append(cal[q].s4.name)
            if list(set(range(date,date+min_shifts+1))&set(doc.days_off)) == [] and list(set(range(date,date+min_shifts))&set(doc.FourSO)) == [] and doc.name not in lastDocs:
                goSched = True
                for d in range(min_shifts):
                    if cal[index + d].s4.name != "--":
                        goSched = False
                """ print([doc.name for doc in gapDocs])
                print(lastDocs)
                print(str(date) + " " + str(gap.remaining) + " " + doc.name) """
                if goSched == True:
                    for j in range(min_shifts):
                        #print(str(index+j-2) + " " + doc.name)
                        cal[index + j].s4 = doc
                        doc.four_shifts += 1
                        gap.remaining -= 1
                        if cal[index + j].weekend == True:
                            doc.weekends += 1
                    gap.docs.append(doc)
                    #schedule additional shift
                    if gap.remaining > 0 and min_shifts != 3 and cal[min(index + min_shifts,len(cal)-1)].s4.name == "--" and not (min_shifts == 2 and (doc.name == "ROD" or doc.name == "TOM")):
                        if date + min_shifts not in doc.days_off and date + min_shifts not in doc.FourSO and date + min_shifts + 1 not in doc.days_off:
                            cal[index + min_shifts].s4 = doc
                            doc.four_shifts += 1
                            gap.remaining -= 1
                            if cal[index + min_shifts].weekend == True:
                                doc.weekends += 1           
                    break

def createGapMap(cal, num_days):
    gapMap = []
    onGap = False
    gapLength = 0
    gapStart = 0
    for i in range(num_days):
        date = i+1
        index = i+3
        if onGap == False and cal[index].s4.name == "--":
            gapStart = date
            gapLength = 1
            onGap = True
        elif onGap == True and cal[index].s4.name == "--":
            gapLength += 1
        elif onGap == True and cal[index].s4.name != "--":
            gapMap.append(Gap(gapStart,gapLength))
            onGap = False
        if date == num_days and onGap == True:
            gapMap.append(Gap(gapStart, gapLength))
    return gapMap

def check4Shifts(docs,cal,year,month):
    docNames = [doc.name for doc in docs]
    printCal(docs,cal,year,month)
    change = input("Would you like to make any 4-shift changes? Type 'Y' or 'N': ").strip().upper()
    while change != "Y" and change != "N":
        change = input("Invalid input. Would you like to make any 4-shift changes? Type 'Y' or 'N': ").strip().upper()
    if change == "N":
        print("\nGreat! I'll schedule the other shifts now, you lazy bum.\n")
    else:
        while True:
            date = input("\nWhich date would you like to change? Type date here: ").strip()
            while True: #get date to change
                if date.isdigit():
                    date = int(date)
                    if date > 0 and date <= calendar.monthrange(year,month)[1]:
                        break
                    else:
                        date = input("Not a valid date. Which date would you like to change? Type date here: ").strip()
                else:
                    date = input("Not a valid date. Which date would you like to change? Type date here: ").strip()
            
            while True: #get new doctor's name
                name = input("Which doctor would you like to take that shift? Type the doctor's name here, the same as it is in the input spreadsheet: ").strip().upper()
                while name not in docNames:
                    name = input("Invalid name. Be sure to spell the doctor's name the same way it is spelled in the input spreadsheet: ").strip().upper()
                index = date + 2
                doc_to_replace = cal[index].s4
                for doc in docs:
                    if doc.name == name:
                        replacement_doc = doc
                while date in replacement_doc.days_off or date in replacement_doc.FourSO or date + 1 in replacement_doc.days_off:
                    inp = input("This causes a scheduling conflict due to days that this doctor requested off. Is this ok? Type 'Y' or 'N': ").strip().upper()
                    if inp == 'Y':
                        break
                    else:
                        name = input("Which doctor would you like to take that shift? Type the doctor's name here, the same as it is in the input spreadsheet: ").strip().upper()
                        while name not in docNames:
                            name = input("Invalid name. Be sure to spell the doctor's name the same way it is spelled in the input spreadsheet: ").strip().upper()
                        for doc in docs:
                            if doc.name == name:
                                replacement_doc = doc


                doc_to_replace.four_shifts -= 1
                replacement_doc.four_shifts += 1
                if cal[index].weekend == True:
                    doc_to_replace.weekends -= 1
                    replacement_doc.weekends += 1
                cal[index].s4 = replacement_doc
                break
            
            printCal(docs,cal,year,month)
            change = input("\nWould you like to make any more 4-shift changes? Type 'Y' or 'N': ").strip().upper()
            while change != "Y" and change != "N":
                change = input("Invalid input. Would you like to make any additional 4-shift changes? Type 'Y' or 'N': ").strip().upper()
            if change == 'N':
                printCal(docs,cal,year,month)
                print("\nGreat! Here's the new schedule. I'll schedule the other shifts now, you lazy bum.\n")
                break

def exportCal(fname,cal,year,month):
    wb = openpyxl.load_workbook(fname)
    sheet = wb["Color"]
    for i in [5,6,7,8,12,13,14,15,19,20,21,22,26,27,28,29,33,34,35,36,41,42,43,44]:
        for j in range(2,9):
            sheet.cell(i,j).value = ""
    firstDay = calendar.monthrange(year,month)[0]
    startRow = 5
    days = 3
    col = firstDay + 3
    if col > 8:
        col -= 7
    while days < len(cal):
        if col > 8:
            col -= 7
            startRow += 7
        sheet.cell(startRow, col).value = cal[days].s1.name
        sheet.cell(startRow + 1,col).value = cal[days].s2.name
        sheet.cell(startRow + 2,col).value = cal[days].s3.name
        sheet.cell(startRow + 3,col).value = cal[days].s4.name
        days += 1
        col += 1
    
    wb.save(fname)

def readInFourShifts(fname,docs,cal,year,month):
    wb = xlrd.open_workbook(fname)
    sheet = wb.sheet_by_name("Color")
    firstDay = calendar.monthrange(year,month)[0]
    startRow = 4
    days = 3
    col = firstDay + 2
    if col > 7:
        col -= 7
    while days < len(cal):
        if col > 7:
            col -= 7
            startRow += 7
        name = sheet.cell_value(startRow + 3, col).strip().upper()
        for doc in docs:
            if doc.name == name:
                cal[days].s4 = doc
        days += 1
        col += 1

def main():
    cal = calendar.TextCalendar()

    inp1 = input("Tu eres garbajo que muerte para dinero. Type 'yes' to agree with this statement: ")
    while inp1.lower().strip() != "yes":
        print()
        print("You absolute nincompoop. Can't you follow one simple instruction? Let's try again.")
        inp1 = input("Tu eres garbajo que muerte para dinero. Type 'yes' to agree with this statement: ")
    print()
    inp1 = input("Glad you agree. Now that that's settled, let's get to scheduling. Press enter to continue: ")
    os.system('cls||clear')

    fname = sys.argv[1]
    docs,month,year,cal1,cal2,cal3 = read_input(fname)
    schedule = []
    #print(cal.prmonth(2020,5))
    #print(calendar.monthrange(2020,8))
    num_days = calendar.monthrange(year,month)[1]
    DOYstart = calendar.monthrange(year,month)[0]
    schedule.append(cal1)
    schedule.append(cal2)
    schedule.append(cal3)
    for i in range(num_days):
        schedule.append(CalDay(i+1))
        weekendTracker = DOYstart + i
        while weekendTracker >= 7:
            weekendTracker -= 7
        if weekendTracker == 5 or weekendTracker == 6:
            schedule[i+3].weekend = True
    buildCal(docs, schedule, num_days, year, month, fname)
    #printCal(docs,schedule,year,month)
    exportCal(fname,schedule,year,month)
    os.system('cls||clear')
    print()
    print("All done, ya filthy animal. Be glad you have a son who is as brilliant as I am. And don't forget: tu eres garbajo que muerte para dinero.")
    print()
    inp = input("Press enter to view the final schedule: ")
    cmd = "start EXCEL.EXE " + fname
    os.system(cmd)
    


if __name__ == "__main__":
    main()
